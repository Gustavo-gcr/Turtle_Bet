import turtle
import time
import random
import openpyxl

class Conta:
    def __init__(self, arquivo_excel):
        self.arquivo_excel = arquivo_excel
        try:
            self.wb = openpyxl.load_workbook(arquivo_excel)
            self.sheet = self.wb.active
            if self.sheet.max_row == 1:
                self.sheet.append(['IDADE', 'NOME', 'SALDO', 'NUM_CONTA'])
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.active
            self.sheet.append(['IDADE', 'NOME', 'SALDO', 'NUM_CONTA'])
            self.wb.save(self.arquivo_excel)
    
    def cadastrar_usuario(self, nome, idade, saldo):
        if idade < 18:
            print(f'Ola, infelizmente sua idade impede de você criar uma conta, já que você tem {idade} anos e a idade mínima exigida é de 18 anos.')
            return False

        numero_conta = random.randint(1, 999)
        nova_conta = {
            'IDADE': idade,
            'NOME': nome,
            'SALDO': saldo,
            'NUM_CONTA': numero_conta
        }
        self.sheet.append([idade, nome, saldo, numero_conta])
        self.wb.save(self.arquivo_excel)
        print(f'Conta cadastrada com sucesso! Número da conta: {numero_conta}')
        return nova_conta

    def depositar(self, numero_conta, valor):
        conta = self.encontrar_conta(numero_conta)
        if conta:
            conta['SALDO'] += valor
            self.atualizar_conta(conta)
            print(f'Depósito de R${valor} realizado com sucesso! Saldo atual: R${conta["SALDO"]}')
        else:
            print('Conta não encontrada.')

    def apostar(self, numero_conta, valor):
        conta = self.encontrar_conta(numero_conta)
        if conta:
            if conta['SALDO'] >= valor:
                conta['SALDO'] -= valor
                self.atualizar_conta(conta)
                print(f'Aposta de R${valor} realizada com sucesso! Saldo atual: R${conta["SALDO"]}')
                return True
            else:
                print(f'Saldo insuficiente para realizar a aposta. Saldo atual: R${conta["SALDO"]}')
                return False
        else:
            print('Conta não encontrada.')
            return False

    def atualizar_saldo(self, numero_conta, valor):
        conta = self.encontrar_conta(numero_conta)
        if conta:
            conta['SALDO'] += valor
            self.atualizar_conta(conta)
            print(f'Saldo atualizado com sucesso! Saldo atual: R${conta["SALDO"]}')
        else:
            print('Conta não encontrada.')

    def atualizar_nome(self, numero_conta, novo_nome):
        conta = self.encontrar_conta(numero_conta)
        if conta:
            conta['NOME'] = novo_nome
            self.atualizar_conta(conta)
            print(f'Nome atualizado com sucesso! Novo nome: {novo_nome}')
        else:
            print('Conta não encontrada.')

    def ver_saldo(self, numero_conta):
        conta = self.encontrar_conta(numero_conta)
        if conta:
            print(f'Saldo da conta {numero_conta}: R${conta["SALDO"]}')
        else:
            print('Conta não encontrada.')

    def ver_todas_contas(self):
        print("\n=== Lista de Contas Cadastradas ===")
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            print(f"Nome: {row[1]}, Número da conta: {row[3]}, Saldo: {'****' if row[2] else '****'}")

    def atualizar_conta(self, conta):
        for row in self.sheet.iter_rows(min_row=2):
            if row[3].value == conta['NUM_CONTA']:
                row[0].value = conta['IDADE']
                row[1].value = conta['NOME']
                row[2].value = conta['SALDO']
                self.wb.save(self.arquivo_excel)
                break

    def encontrar_conta(self, numero_conta):
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[3] == numero_conta:
                return {
                    'IDADE': row[0],
                    'NOME': row[1],
                    'SALDO': row[2],
                    'NUM_CONTA': row[3]
                }
        return None

def exibir_menu():
    print("\n=== MENU ===")
    print("1. Fazer aposta")
    print("2. Cadastrar uma nova conta")
    print("3. Depositar Dinheiro")
    print("4. Ver saldo da conta")
    print("5. Atualizar nome do titular")
    print("6. Ver todas as contas cadastradas")
    print("7. Sair do programa")

WIDTH, HEIGHT = 700, 600
CORES = ['red', 'green', 'blue', 'orange', 'yellow', 'black', 'purple', 'pink', 'brown', 'cyan']

def numero_tartaruga():
    racers = 0
    while True:
        racers = input('Escolha o numero de competidoras (2 - 10): ')
        if racers.isdigit():
            racers = int(racers)
        else:
            print('Digite um número, tente novamente')
            continue

        if 2 <= racers <= 10:
            return racers
        else:
            print('Número não está entre 2 e 10')

def corrida(cores):
    turtles = criar_tartarugas(cores)

    while True:
        for racer in turtles:
            distance = random.randrange(1, 20)
            racer.forward(distance)

            x, y = racer.pos()
            if y >= HEIGHT // 2 - 10:
                return cores[turtles.index(racer)]

def criar_tartarugas(cores):
    turtles = []
    spacingx = WIDTH // (len(cores) + 1)
    for i, color in enumerate(cores):
        racer = turtle.Turtle()
        racer.color(color)
        racer.shape('turtle')
        racer.left(90)
        racer.penup()
        racer.setpos(-WIDTH // 2 + (i + 1) * spacingx, -HEIGHT // 2 + 20)
        racer.pendown()
        turtles.append(racer)

    return turtles

def init_turtle():
    screen = turtle.Screen()
    screen.setup(WIDTH, HEIGHT)
    screen.title('Turtle Racing!')
    return screen

def main():
    conta = Conta('contas.xlsx')

    while True:
        exibir_menu()
        opcao = input('Escolha uma opção: ')

        if opcao == '1':
            numero_conta = int(input('Número da conta: '))
            valor_aposta = float(input('Valor da aposta: R$'))

            if conta.apostar(numero_conta, valor_aposta):
                racers = numero_tartaruga()

                random.shuffle(CORES)
                cores = CORES[:racers]

                escolha = input(f'Escolha uma tartaruga para apostar {cores}: ').strip().lower()

                screen = init_turtle()
                winner = corrida(cores)

                if escolha == winner:
                    if racers == 2:
                        premio = valor_aposta * 1.2
                    elif racers == 3:
                        premio = valor_aposta * 1.6
                    elif racers == 4:
                        premio = valor_aposta * 1.75
                    elif racers == 5:
                        premio = valor_aposta * 1.86
                    elif racers == 6:
                        premio = valor_aposta * 2
                    elif racers == 7:
                        premio = valor_aposta * 3
                    elif racers == 8:
                        premio = valor_aposta * 4.7
                    elif racers == 9:
                        premio = valor_aposta * 5
                    else:  # racers = 10
                        premio = valor_aposta * 7.7

                    print(f'Você ganhou! O prêmio é R${premio}')
                    conta.atualizar_saldo(numero_conta, premio)
                else:
                    print(f'Você perdeu! A tartaruga vencedora foi a da cor {winner}')
                
                time.sleep(5)
                screen.bye()
            else:
                print('Aposta não realizada. Saldo insuficiente.')

        elif opcao == '2':
            nome = input('Nome do titular: ')
            idade = int(input('Idade: '))
            saldo = float(input('Saldo inicial: R$'))
            conta.cadastrar_usuario(nome, idade, saldo)

        elif opcao == '3':
            numero_conta = int(input('Número da conta: '))
            valor = float(input('Valor para depósito: R$'))
            conta.depositar(numero_conta, valor)

        elif opcao == '4':
            numero_conta = int(input('Número da conta: '))
            conta.ver_saldo(numero_conta)

        elif opcao == '5':
            numero_conta = int(input('Número da conta: '))
            novo_nome = input('Novo nome do titular: ')
            conta.atualizar_nome(numero_conta, novo_nome)

        elif opcao == '6':
            conta.ver_todas_contas()

        elif opcao == '7':
            print('Programa encerrado. Obrigado por utilizar!')
            break

        else:
            print('Opção inválida. Escolha novamente.')

if __name__ == "__main__":
    main()
